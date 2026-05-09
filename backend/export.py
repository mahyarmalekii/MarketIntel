"""
Export portfolio data to Excel (.xlsx) and Google Sheets.
"""
import os
import tempfile
from datetime import datetime


def export_to_excel(
    holdings: list[dict],
    news: list[dict],
    insights: list[dict],
    scenarios: list[dict],
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    _hdr_font = Font(bold=True, color="FFFFFF", size=11)
    _hdr_fill = PatternFill("solid", fgColor="1F1A14")
    _green_fill = PatternFill("solid", fgColor="DCEAD4")
    _red_fill = PatternFill("solid", fgColor="F6DDE3")
    _border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def _hdr_row(ws, cols: list[str]):
        ws.append(cols)
        for cell in ws[ws.max_row]:
            cell.font = _hdr_font
            cell.fill = _hdr_fill
            cell.alignment = Alignment(horizontal="center")

    def _col_widths(ws, widths: list[int]):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # ── Portfolio sheet ──
    ws1 = wb.active
    ws1.title = "Portfolio"
    _hdr_row(ws1, ["Ticker", "Name", "Region", "Shares", "Avg Cost", "Current Price",
                    "Value", "Cost Basis", "P&L", "P&L %", "Sector", "Cap Tier", "Notes"])
    _col_widths(ws1, [10, 22, 8, 10, 12, 14, 14, 14, 12, 10, 16, 10, 20])
    total_value = total_cost = total_pnl = 0.0
    for h in holdings:
        pnl = h.get("pnl", 0) or 0
        pnl_pct = h.get("pnl_pct", 0) or 0
        val = h.get("current_value", 0) or 0
        cost = h.get("cost_basis", 0) or 0
        total_value += val
        total_cost += cost
        total_pnl += pnl
        row = ws1.append([
            h.get("ticker", ""), h.get("name", ""), h.get("region", ""),
            h.get("shares", 0), h.get("avg_cost", 0), h.get("current_price", 0),
            round(val, 2), round(cost, 2), round(pnl, 2),
            f"{pnl_pct:+.2f}%",
            h.get("sector", ""), h.get("market_cap_tier", ""), h.get("notes", ""),
        ])
        # color P&L cell
        pnl_cell = ws1.cell(ws1.max_row, 9)
        pnl_cell.fill = _green_fill if pnl >= 0 else _red_fill

    # Summary row
    ws1.append([])
    ws1.append(["TOTAL", "", "", "", "", "",
                round(total_value, 2), round(total_cost, 2), round(total_pnl, 2),
                f"{(total_pnl/total_cost*100 if total_cost else 0):+.2f}%"])
    for cell in ws1[ws1.max_row]:
        cell.font = Font(bold=True)

    # ── News sheet ──
    ws2 = wb.create_sheet("News")
    _hdr_row(ws2, ["Title", "Source", "Sentiment", "Tickers", "Published", "URL"])
    _col_widths(ws2, [50, 20, 12, 20, 20, 50])
    for n in news:
        ws2.append([
            n.get("title", ""), n.get("source", ""), n.get("sentiment", ""),
            ", ".join(n.get("tickers", [])), n.get("published_at", ""), n.get("url", ""),
        ])
        sent = n.get("sentiment", "")
        cell = ws2.cell(ws2.max_row, 3)
        if sent == "bullish":
            cell.fill = _green_fill
        elif sent == "bearish":
            cell.fill = _red_fill

    # ── Insights sheet ──
    ws3 = wb.create_sheet("Insights")
    _hdr_row(ws3, ["Ticker", "Action", "Confidence %", "Target Price", "Scenario",
                    "Rationale", "Model", "Created"])
    _col_widths(ws3, [10, 8, 14, 14, 10, 60, 18, 20])
    for ins in insights:
        ws3.append([
            ins.get("ticker", ""), ins.get("action", ""), ins.get("confidence", 0),
            ins.get("target_price", ""), ins.get("scenario", ""),
            ins.get("rationale", ""), ins.get("model_used", ""), ins.get("created_at", ""),
        ])
        action = ins.get("action", "")
        cell = ws3.cell(ws3.max_row, 2)
        if action == "buy":
            cell.fill = _green_fill
        elif action == "sell":
            cell.fill = _red_fill

    # ── Scenarios sheet ──
    ws4 = wb.create_sheet("Scenarios")
    _hdr_row(ws4, ["Name", "Portfolio Impact %", "Affected Sectors", "Assumptions",
                    "Description", "Created"])
    _col_widths(ws4, [20, 18, 30, 40, 40, 20])
    for sc in scenarios:
        ws4.append([
            sc.get("name", ""),
            f"{sc.get('portfolio_impact', 0) or 0:+.1f}%",
            ", ".join(sc.get("affected_sectors", [])),
            sc.get("assumptions", ""), sc.get("description", ""), sc.get("created_at", ""),
        ])
        impact = sc.get("portfolio_impact") or 0
        cell = ws4.cell(ws4.max_row, 2)
        cell.fill = _green_fill if impact >= 0 else _red_fill

    path = os.path.join(
        tempfile.gettempdir(),
        f"portfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    wb.save(path)
    return path


def export_to_gsheets(
    sheet_id: str,
    holdings: list[dict],
    insights: list[dict],
    credentials_path: str = "",
) -> None:
    import gspread

    if credentials_path and os.path.exists(credentials_path):
        gc = gspread.service_account(filename=credentials_path)
    else:
        gc = gspread.service_account()

    sh = gc.open_by_key(sheet_id)

    # Portfolio tab
    try:
        ws = sh.worksheet("Portfolio")
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet("Portfolio", rows=200, cols=13)

    headers = ["Ticker", "Region", "Shares", "Avg Cost", "Current Price",
               "Value", "P&L", "P&L %", "Sector"]
    rows = [headers]
    for h in holdings:
        rows.append([
            h.get("ticker", ""), h.get("region", ""), h.get("shares", 0),
            h.get("avg_cost", 0), h.get("current_price", 0),
            round(h.get("current_value", 0) or 0, 2),
            round(h.get("pnl", 0) or 0, 2),
            f"{h.get('pnl_pct', 0) or 0:+.2f}%",
            h.get("sector", ""),
        ])
    ws.update(rows)

    # Insights tab
    try:
        wi = sh.worksheet("Insights")
        wi.clear()
    except gspread.WorksheetNotFound:
        wi = sh.add_worksheet("Insights", rows=200, cols=7)

    i_headers = ["Ticker", "Action", "Confidence %", "Target Price", "Scenario",
                  "Rationale", "Created"]
    i_rows = [i_headers]
    for ins in insights:
        i_rows.append([
            ins.get("ticker", ""), ins.get("action", ""), ins.get("confidence", 0),
            ins.get("target_price", ""), ins.get("scenario", ""),
            ins.get("rationale", ""), ins.get("created_at", ""),
        ])
    wi.update(i_rows)
